# coding:utf-8
'''
Nils Napp, Yitong Sun, Xu Hai Nov 2021
Script to group and count the entry/exit events form beetunnel filenames
'''

import cv2
import pathlib as path
import datetime as dt
import numpy as np
import time
import webbrowser
import sys
import os
from matplotlib import pyplot as plt
import bee_event as bee  # be sure to include bee_event.py in the same dir
from sys import platform
import re
import pickle
import math
import xlsxwriter
import openpyxl

# Time difference for grouping images
DELTA_T_SEC = 2

FMT_STR = '%Y-%m-%d_%H-%M-%S_%f'
MIN_PIC = 4

start_time = time.time()


def tstr(fname):
    parts = fname.split('_')
    return parts[0] + '_' + parts[1] + '_' + parts[2]


def delta_dt_sec(dt1, dt2):
    delta = dt2 - dt1
    return delta.days * 24 * 60 * 60 + delta.seconds + delta.microseconds / 1000000


def get_events(DIR_REL):
    p = path.Path(DIR_REL)
    dtimes = [dt.datetime.strptime(tstr(f.name), FMT_STR) for f in p.glob('*.jpg')]
    dtimes.sort()
    if len(dtimes) <= 2:
        return None

    events = [0]
    event_delta = []
    dt0 = dtimes[0]
    dt_last = dt0
    pic_per_ev = []
    pic_cnt = 0

    dt_list = []

    events_list = []
    event_elements = []
    for d in dtimes:
        pic_cnt += 1
        ev_del = delta_dt_sec(dt_last, d)
        if ev_del > DELTA_T_SEC:
            if pic_cnt < MIN_PIC:
                pic_cnt = 1
            else:
                pic_per_ev.append(pic_cnt)
                pic_cnt = 1
                events.append(delta_dt_sec(dt0, d))
                event_delta.append(ev_del)
                events_list.append(event_elements)
                event_elements = []
        else:
            # print(ev_del)
            event_elements.append(d)
            if ev_del < .2:
                dt_list.append(ev_del)

        dt_last = d

    ret = {}
    ret['events'] = events
    ret['event_delta'] = event_delta
    ret['event_cnt'] = pic_per_ev
    ret['start_dt'] = dtimes[0]
    ret['ev_dt'] = dt_list
    ret['events_list'] = events_list
    return ret


# Get Image V3
def getImage(d, f_dir_d):
    d_name = ""

    year = str(d.year)
    if d.month < 10:
        month = "0" + str(d.month)
    else:
        month = str(d.month)
    if d.day < 10:
        day = "0" + str(d.day)
    else:
        day = str(d.day)
    if d.hour < 10:
        hour = "0" + str(d.hour)
    else:
        hour = str(d.hour)
    if d.minute < 10:
        minus = "0" + str(d.minute)
    else:
        minus = str(d.minute)
    if d.second < 10:
        sec = "0" + str(d.second)
    else:
        sec = str(d.second)
    if d.microsecond / 10 < 1:
        ms = "00000" + str(d.microsecond)
    elif d.microsecond / 100 < 1:
        ms = "0000" + str(d.microsecond)
    elif d.microsecond / 1000 < 1:
        ms = "000" + str(d.microsecond)
    elif d.microsecond / 10000 < 1:
        ms = "00" + str(d.microsecond)
    elif d.microsecond / 100000 < 1:
        ms = "0" + str(d.microsecond)
    else:
        ms = str(d.microsecond)
    name_d = year + "-" + month + "-" + day + "_" + hour + "-" + minus + "-" + sec + "_" + ms + "*"

    p_d = path.Path(f_dir_d)

    for f in p_d.glob(name_d):
        d_name = str(f)

    if d_name != "":
        return d_name
    else:
        print('Unexpected error: previous file can not be found')
        print('File name: ' + name_d)
        quit()


# Match same event for top and side views
def event_match(es_list_a, es_list_b):
    event_matched = {}
    visited = []
    # event_start_time = str(int(in_date_parts[1])) + str(int(in_date_parts[2])) + str(int(in_time_parts[0]))
    for i in range(len(es_list_a)):
        event_start_time = str(es_list_a[i][0].day) + str(es_list_a[i][0].hour) + str(es_list_a[i][0].minute)
        for j in range(len(es_list_b)):
            event_b_start_time = str(es_list_b[j][0].day) + str(es_list_b[j][0].hour) + str(es_list_b[j][0].minute)
            if event_b_start_time == event_start_time:
                if abs(es_list_b[j][0].second - es_list_a[i][0].second) <= 2 and j not in visited:
                    visited.append(j)
                    event_matched[i] = j

    # if len(event_a) != len(event_b):
    #     print('Some images of top/side view misses, only take the first satisfied event.')
    #     event_a = [event_a[0]]
    #     event_b = [event_b[0]]
    return event_matched


pattern_date = re.compile(r'(\d{4}-\d{2}-\d{2})')
pattern_time = re.compile(r'(\d{2}-\d{2}-\d{2})')

# Let the user input date and time
# And make sure the format is correct
# count = 0
# while 1:
#     in_date = input("Please input a date you want to query (follow the format 'yyyy-mm-dd'): ")
#     if re.match(pattern_date, in_date):
#         break
#     count += 1
#     if count == 5:
#         print('Too many errors, program stops.')
#         quit()
#
# count = 0
#
# while 1:
#     in_time = input("Please input the time you want to query (follow the format 'hh-mm-ss'): ")
#     if re.match(pattern_time, in_time):
#         break
#     count += 1
#     if count == 5:
#         print('Too many errors, program stops.')
#         quit()

wrkbk = openpyxl.load_workbook("Pollen_Project_new.xlsx")
sh = wrkbk.active
count = 0
names = []

pollen_backup = xlsxwriter.Workbook('pollen_backup.xlsx')

# The workbook object is then used to add new
# worksheet via the add_worksheet() method.
pollen_sheet = pollen_backup.add_worksheet()

# Use the worksheet object to write
# data via the write() method.
pollen_sheet.write('A1', 'Sample')
pollen_sheet.write('B1', 'Time')
pollen_sheet.write('C1', 'Bee ID')

final_pickle_file = {}

for i in range(65, 142):#sh.max_row + 1):
    pre_date = str(sh.cell(row=i, column=1).value)
    pre_time = str(sh.cell(row=i, column=2).value)
    bee_id = str(sh.cell(row=i, column=3).value)

    in_date = pre_date.split(' ')[1].split('/')
    in_time = pre_time.split(':')

    in_date_parts = [0 for x in range(3)]
    in_time_parts = [0 for x in range(3)]

    in_date_parts[0] = in_date[2]
    if len(in_date[0]) < 2:
        in_date_parts[1] = "0" + in_date[0]
    else:
        in_date_parts[1] = in_date[0]
    if len(in_date[1]) < 2:
        in_date_parts[2] = "0" + in_date[1]
    else:
        in_date_parts[2] = in_date[1]

    if len(in_time[0]) < 2:
        in_time_parts[0] = "0" + in_time[0]
    else:
        in_time_parts[0] = in_time[0]
    if len(in_time[1]) < 2:
        in_time_parts[1] = "0" + in_time[1]
    else:
        in_time_parts[1] = in_time[1]
    in_time_parts[2] = "59"

    top_files = os.listdir('/Volumes/BEE_DATA/run_top-4___2021-09-26_18-00-44/')
    side_files = os.listdir('/Volumes/BEE_DATA/run_side-4___2021-09-26_18-00-51/')

    latest_time_top = 0
    latest_time_side = 0

    pre_count = count

    for f in top_files:
        parts = f.split('_')
        if len(parts) == 2:
            tmp_month = parts[0].split('-')[1]
            tmp_day = parts[0].split('-')[2]
            tmp_hour = int(parts[1].split('-')[0])
            tmp_minute = int(parts[1].split('-')[1])
            tmp_second = int(parts[1].split('-')[2])

            if tmp_month == in_date_parts[1]:
                if in_date_parts[2] == tmp_day:
                    if int(in_time_parts[0]) * 60 * 60 + int(in_time_parts[1]) * 60 + int(in_time_parts[2]) >= \
                            tmp_hour * 60 * 60 + tmp_minute * 60 + tmp_second > latest_time_top:
                        latest_time_top = tmp_hour * 60 * 60 + tmp_minute * 60 + tmp_second
                        aim_file_top = f

    for f in side_files:
        parts = f.split('_')
        if len(parts) == 2:
            tmp_month = parts[0].split('-')[1]
            tmp_day = parts[0].split('-')[2]
            tmp_hour = int(parts[1].split('-')[0])
            tmp_minute = int(parts[1].split('-')[1])
            tmp_second = int(parts[1].split('-')[2])

            if tmp_month == in_date_parts[1]:
                if in_date_parts[2] == tmp_day:
                    if int(in_time_parts[0]) * 60 * 60 + int(in_time_parts[1]) * 60 + int(in_time_parts[2]) >= \
                            tmp_hour * 60 * 60 + tmp_minute * 60 + tmp_second > latest_time_side:
                        latest_time_side = tmp_hour * 60 * 60 + tmp_minute * 60 + tmp_second
                        aim_file_side = f

    try:
        aim_file_top = '/Volumes/BEE_DATA/run_top-4___2021-09-26_18-00-44/' + str(aim_file_top) + '/'
        aim_file_side = '/Volumes/BEE_DATA/run_side-4___2021-09-26_18-00-51/' + str(aim_file_side) + '/'
    except NameError:
        print('No data can be found with input date and time.')
        quit()

    r3 = get_events(aim_file_top)
    r3_side = get_events(aim_file_side)

    es_list = r3.get('events_list')
    nums = r3.get('event_cnt')

    es_list_2 = r3_side.get('events_list')
    nums_2 = r3_side.get('event_cnt')

    events_matched = event_match(es_list, es_list_2)

    pickle_file_top = {}
    pickle_file_side = {}

    len_count = 0

    for n in events_matched:
        pickle_file = {}
        top_index = n
        side_index = events_matched[n]

        # All images of an event
        # The first image is the middle image of the event
        top_events = []
        side_events = []

        # The middle image of an event
        top_mid = es_list[top_index][int(len(es_list[top_index]) / 2)]
        side_mid = es_list_2[side_index][int(len(es_list_2[side_index]) / 2)]

        # Record the middle image name
        top_events.append(getImage(top_mid, aim_file_top))
        side_events.append(getImage(side_mid, aim_file_side))

        # Record all images of an event
        for e in es_list[top_index]:
            top_events.append(getImage(e, aim_file_top))

        for e in es_list_2[side_index]:
            side_events.append(getImage(e, aim_file_side))

        res_list = []

        names_flag = 0

        if platform == "win32":
            names.append(str(top_events[0].split('\\')[5].split('_')[0]) + '-' +
                         str(top_events[0].split('\\')[5].split('_')[1]) + '.html')
            names_flag = 1
        else:
            names.append(str(top_events[0].split('/')[5].split('_')[0]) + '-' +
                         str(top_events[0].split('/')[5].split('_')[1]) + '.html')
            names_flag = 1

        if names_flag == 0:
            continue

        top_values = []
        side_values = []

        for image in top_events:
            top_values.append(image[17:])
        for image in side_events:
            side_values.append(image[17:])
        # pickle_file_top[names[-1][:-5]] = top_values
        # pickle_file_side[names[-1][:-5]] = side_values
        pickle_file_top[names[count][:-5]] = top_values
        pickle_file_side[names[count][:-5]] = side_values

        pickle_file["top"] = top_values
        pickle_file["side"] = side_values

        final_pickle_file[names[count][:-5]] = pickle_file
        # final_pickle_file[names[-1][:-5]] = pickle_file

        # Previous button
        if count > 0:
            res_list.append("<a href='file:" + str(count) + ".html")
            res_list.append("' class='previous'> &laquo Previous </a>")
        # Next button
        if i != 141 or len_count < len(events_matched) - 1:
            res_list.append("<a href='file:" + str(count + 2) + ".html")
            res_list.append("' class='next'>Next &raquo;</a>")

        len_count += 1
        # Record the title of each event
        res_list.append("<div class='event'>")
        res_list.append("<p class='event_num'>Event " + names[count][: -5] + "</p>")

        count += 1

        res_list.append("<div class='area_block'>")
        # Deal with different OS
        # Windows
        if platform == "win32":
            res_list.append("<p class='event_des'>Date: " + str(top_events[0].split('\\')[5].split('_')[0]) + "</p>")
            res_list.append("<p class='event_des'>Time: " + str(top_events[0].split('\\')[5].split('_')[1]) + "</p>")
        # Linux / Mac
        else:
            res_list.append("<p class='event_des'>Date: " + str(top_events[0].split('/')[5].split('_')[0]) + "</p>")
            res_list.append("<p class='event_des'>Time: " + str(top_events[0].split('/')[5].split('_')[1]) + "</p>")
        res_list.append("</div>")

        res_list.append("<div class='area_block'>")
        res_list.append("<p class='view_dir'>Side View</p>")
        res_list.append("<img src=" + str(side_events[0]) + " class='img' height='200' width='300'>")  # sid image
        res_list.append("</div>")

        res_list.append("<div class='area_block'>")
        res_list.append("<p class='view_dir'>Top View</p>")
        res_list.append("<img src=" + str(top_events[0]) + " class='img' height='200' width='300'>")  # top image
        res_list.append("</div>")

        res_list.append("</div>")

        # Record all images of the side view
        res_list.append("<div class='event_view'>")
        res_list.append("<p class='event_num'>All Images of the Side View</p>")

        res_list.append("<div class='event_imgs'>")

        for j in range(1, len(side_events)):
            if j % 9 == 0:
                res_list.append("<div>")
            res_list.append("<div class='area_block'>")
            res_list.append("<img src=" + str(side_events[j]) + " class='img' height='50' width='80'>")
            res_list.append("</div>")
            if j % 9 == 0:
                res_list.append("</div>")

        res_list.append("</div>")

        res_list.append("</div>")

        # Record all images of the top view
        res_list.append("<div class='event_view'>")
        res_list.append("<p class='event_num'>All Images of the Top View</p>")

        res_list.append("<div class='event_imgs'>")

        for j in range(1, len(top_events)):
            if j % 9 == 0:
                res_list.append("<div>")
            res_list.append("<div class='area_block'>")
            res_list.append("<img src=" + str(top_events[j]) + " class='img' height='50' width='80'>")
            res_list.append("</div>")
            if j % 9 == 0:
                res_list.append("</div>")

        res_list.append("</div>")

        res_list.append("</div>")

        img_str = ""
        for tags in res_list:
            img_str += tags

        # html generated
        if platform == "win32":
            GEN_HTML = str(count) + ".html" #names[-1]
        else:
            GEN_HTML = str(count) + ".html" #names[-1]

        # open html
        f = open(GEN_HTML, 'w', encoding="utf-8")

        message = """
        <!DOCTYPE html>
        <html>
        <head>
            <link rel="stylesheet" type="text/css" href="task001.css">
            <meta charset="utf-8">
            <title>test</title>
            <style type="text/css">
            *{margin: 0; padding: 0;}
            .clearfix:after {
                  clear: both;
                  content: ".";
                  display: block;
                  height: 0;
                  visibility: hidden;
                }
                .event_imgs {
                    margin: auto;
                    height: 300px;
                }
                .area_block {
                    float: left;
                    margin: 20px;
                }
                .view_dir {
                    font-size: 18px;
                    font-weight: 600;
                    color: black;
                }
                .event_des {
                    font-size: 16px;
                    font-weight: 500;
                    color: black;
                    float: left;
                    margin: 0 20px 0;
                }
                .event {
                    margin-bottom: 50px;
                    height: 300px;
                }
                .event_view {
                    margin: 20px 0px 20px;
                    height: 350px;
                }
                .wrap {
                    padding: 20px;
                    border-radius: 25px;
                    width: 1200px;
                    text-align: center;
                    margin: 0 auto;
                    background: rgb(241, 239, 239);
                    box-shadow: 0 12px 16px 0 rgba(0,0,0,0.24), 0 17px 50px 0 rgba(0,0,0,0.19);
                }
                #header {
                    background-color:rgb(160, 39, 39);
                    color:white;
                    text-align:center;
                    padding:5px;
                    border: none;
                    box-shadow: 0 12px 16px 0 rgba(0,0,0,0.24), 0 17px 50px 0 rgba(0,0,0,0.19);
                    margin-bottom: 90px;
                }
                body {
                    background: rgb(207, 204, 204);
                    margin: 0;
                }
                .event_num {
                    font-size: 28px;
                    font-weight: 900;
                    color: black;
                    margin-top: 10px;
                    margin-bottom: 10px;
                }
                #footer {
                    background-color:rgb(160, 39, 39);
                    color:white;
                    clear:both;
                    text-align:center;
                    padding:5px;
                }
                a {
                text-decoration: none;
                display: inline-block;
                padding: 8px 16px;
                }
                a:hover {
                background-color: #ddd;
                color: black;
                }
                .previous {
                background-color: #f1f1f1;
                color: black;
                }
                .next {
                background-color: #04AA6D;
                color: white;
                }
            </style>
        </head>
        <body>
            <div id="header">
            <h1>Custom Bee Tracking System</h1>
            </div>
            <div class="wrap">
                %s
            </div>
        </body>
        </html>""" % (img_str)

        f.write(message)

        f.close()

        # run web broswer
        # webbrowser.open(GEN_HTML, new=1)
        '''
        webbrowser.open(url, new=0, autoraise=True)
        Display url using the default browser. If new is 0, the url is opened in the same browser window if possible. If new is 1, a new browser window is opened if possible. If new is 2, a new browser page (“tab”) is opened if possible. If autoraise is True, the window is raised if possible (note that under many window managers this will occur regardless of the setting of this variable).
        '''
    #webbrowser.open_new_tab('file:' + os.path.realpath(str(count) + ".html"))

    cnt = 0
    # for j in events_matched:
    #     print("===========================")
    #     print(cnt)
    #     cnt += 1
    #     print("Top event")
    #     print(es_list[j])
    #     print("Side")
    #     print(es_list_2[events_matched[j]])

    file_name = names[-1]
    outfile_top = open(file_name + "_top", 'wb')
    pickle.dump(pickle_file_top, outfile_top)
    outfile_top.close()

    outfile_side = open(file_name + "_side", 'wb')
    pickle.dump(pickle_file_side, outfile_side)
    outfile_side.close()

    # Workbook() takes one, non-optional, argument
    # which is the filename that we want to create.
    workbook = xlsxwriter.Workbook(file_name + '.xlsx')

    # The workbook object is then used to add new
    # worksheet via the add_worksheet() method.
    worksheet = workbook.add_worksheet()

    # Use the worksheet object to write
    # data via the write() method.
    worksheet.write('A1', 'Event')
    worksheet.write('C1', 'Bee ID')
    worksheet.write('D1', 'Pollen')
    worksheet.write('E1', 'Top')
    worksheet.write('P1', 'Side')

    cnt = 2

    # Match corresponding top view and side view images
    # And merge them into vertical format
    for a in pickle_file_top:
        min_len = min(len(pickle_file_side[a]), len(pickle_file_top[a]))
        worksheet.write_url('A' + str(cnt), r'./' + str(pre_count + 1) + '.html', string=a)
        for j in range(min_len):
            worksheet.write('E' + str(cnt), pickle_file_top[a][j])
            worksheet.write('P' + str(cnt), pickle_file_side[a][j])
            cnt += 1
        pre_count += 1

    # Finally, close the Excel file
    # via the close() method.
    workbook.close()

    pollen_sheet.write_url('A' + str(i - 64), r'./' + file_name + '.xlsx', string=pre_date)
    pollen_sheet.write('B' + str(i - 64), pre_time)
    pollen_sheet.write('C' + str(i - 64), bee_id)
pollen_backup.close()

outfile_a = open("all_events", 'wb')
pickle.dump(final_pickle_file, outfile_a)
outfile_a.close()

while 1:
    delete_or_not = input("Do you want to remove generated HTML files? (y/n): ")
    if delete_or_not.lower() == 'y' or delete_or_not.lower() == 'yes':
        for i in range(1, count + 1):
            os.remove(str(i) + '.html')
        for i in names:
            try:
                os.remove(i + '_top')
                os.remove(i + '_side')
            except FileNotFoundError:
                continue
        break
    elif delete_or_not.lower() == 'n' or delete_or_not.lower() == 'no':
        break


